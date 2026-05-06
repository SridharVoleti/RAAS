from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from domain import new_id


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def import_courses_from_xlsx(xlsx_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(filename=BytesIO(xlsx_bytes))
    ws = wb.active

    header = [(_cell_str(c.value)).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def idx(name: str) -> Optional[int]:
        name = name.lower()
        if name in header:
            return header.index(name)
        return None

    course_title_i = idx("course_title")
    course_desc_i = idx("course_description")
    course_price_i = idx("course_price")
    lesson_title_i = idx("lesson_title")
    youtube_id_i = idx("youtube_id")

    learning_path_title_i = idx("learning_path_title")
    learning_path_desc_i = idx("learning_path_description")
    learning_path_price_i = idx("learning_path_price")

    if course_title_i is None or lesson_title_i is None or youtube_id_i is None:
        raise ValueError("missing_required_columns")

    courses_by_title: Dict[str, Dict[str, Any]] = {}
    learning_paths_by_title: Dict[str, Dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=2):
        course_title = _cell_str(row[course_title_i].value)
        if not course_title:
            continue

        course = courses_by_title.get(course_title)
        if not course:
            course = {
                "id": new_id("crs"),
                "title": course_title,
                "description": _cell_str(row[course_desc_i].value) if course_desc_i is not None else "",
                "price": float(row[course_price_i].value) if course_price_i is not None and row[course_price_i].value is not None else 0.0,
                "lessons": [],
            }
            courses_by_title[course_title] = course

        lesson_title = _cell_str(row[lesson_title_i].value)
        youtube_id = _cell_str(row[youtube_id_i].value)
        if not lesson_title or not youtube_id:
            continue

        course["lessons"].append(
            {
                "id": new_id("les"),
                "title": lesson_title,
                "youtube_id": youtube_id,
            }
        )

        if learning_path_title_i is not None:
            learning_path_title = _cell_str(row[learning_path_title_i].value)
            if learning_path_title:
                lp = learning_paths_by_title.get(learning_path_title)
                if not lp:
                    lp = {
                        "id": new_id("lp"),
                        "title": learning_path_title,
                        "description": _cell_str(row[learning_path_desc_i].value) if learning_path_desc_i is not None else "",
                        "price": float(row[learning_path_price_i].value) if learning_path_price_i is not None and row[learning_path_price_i].value is not None else 0.0,
                        "course_ids": [],
                    }
                    learning_paths_by_title[learning_path_title] = lp

                if course.get("id") not in lp["course_ids"]:
                    lp["course_ids"].append(course.get("id"))

    return {"courses": list(courses_by_title.values()), "learning_paths": list(learning_paths_by_title.values())}
