import pytest

from app import create_app


@pytest.fixture()
def client(tmp_path):
    app = create_app()
    app.config["TESTING"] = True
    app.config["DATA_DIR"] = str(tmp_path)

    with app.test_client() as c:
        yield c


def _auth_headers(token: str):
    return {"Authorization": f"Bearer {token}"}


def test_progress_requires_95_percent(client, tmp_path):
    admin_login = client.post("/api/auth/login", json={"email": "sridhar.voleti@gmail.com", "password": "admin123"}).get_json()
    admin_token = admin_login["token"]

    # Import a minimal course
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["course_title", "lesson_title", "youtube_id"])
    ws.append(["Course 1", "L1", "abc123"])
    ws.append(["Course 1", "L2", "def456"])

    buf = BytesIO()
    wb.save(buf)

    r_import = client.post(
        "/api/admin/import/courses",
        headers=_auth_headers(admin_token),
        data={"file": (BytesIO(buf.getvalue()), "courses.xlsx")},
        content_type="multipart/form-data",
    )
    assert r_import.status_code == 200

    reg = client.post("/api/auth/register", json={"email": "s@t.com", "password": "pw"}).get_json()
    token = reg["token"]

    courses = client.get("/api/learner/courses").get_json()["courses"]
    course_id = courses[0]["id"]

    enr = client.post("/api/learner/enroll", json={"course_id": course_id}, headers=_auth_headers(token)).get_json()["enrollment"]

    # Admin approves enrollment
    all_enr = client.get("/api/admin/enrollments", headers=_auth_headers(admin_token)).get_json()["enrollments"]
    approve_id = next(e["id"] for e in all_enr if e["user_id"] == reg["user"]["id"]) 
    r_app = client.post(f"/api/admin/enrollments/{approve_id}/approve", headers=_auth_headers(admin_token))
    assert r_app.status_code == 200

    outline = client.get(f"/api/learner/course/{course_id}/outline", headers=_auth_headers(token)).get_json()
    lessons = outline["lessons"]
    assert lessons[0]["enabled"] is True
    assert lessons[1]["enabled"] is False

    token1 = client.post(
        f"/api/learner/course/{course_id}/playback-token",
        json={"lesson_id": lessons[0]["id"]},
        headers=_auth_headers(token),
    ).get_json()["token"]

    # 94% watched should not unlock next
    r_hb = client.post(
        f"/api/learner/course/{course_id}/heartbeat",
        json={"playback_token": token1, "position_seconds": 94.0, "duration_seconds": 100.0},
        headers=_auth_headers(token),
    )
    assert r_hb.status_code == 200

    outline2 = client.get(f"/api/learner/course/{course_id}/outline", headers=_auth_headers(token)).get_json()
    assert outline2["lessons"][1]["enabled"] is False

    # 95% watched unlocks next
    r_hb2 = client.post(
        f"/api/learner/course/{course_id}/heartbeat",
        json={"playback_token": token1, "position_seconds": 95.0, "duration_seconds": 100.0},
        headers=_auth_headers(token),
    )
    assert r_hb2.status_code == 200

    outline3 = client.get(f"/api/learner/course/{course_id}/outline", headers=_auth_headers(token)).get_json()
    assert outline3["lessons"][1]["enabled"] is True


def test_learning_path_gating_across_courses(client):
    admin_login = client.post("/api/auth/login", json={"email": "sridhar.voleti@gmail.com", "password": "admin123"}).get_json()
    admin_token = admin_login["token"]

    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["learning_path_title", "course_title", "lesson_title", "youtube_id"])
    ws.append(["Path A", "Course 1", "C1-L1", "abc123"])
    ws.append(["Path A", "Course 2", "C2-L1", "def456"])

    buf = BytesIO()
    wb.save(buf)

    r_import = client.post(
        "/api/admin/import/courses",
        headers=_auth_headers(admin_token),
        data={"file": (BytesIO(buf.getvalue()), "courses.xlsx")},
        content_type="multipart/form-data",
    )
    assert r_import.status_code == 200

    reg = client.post("/api/auth/register", json={"email": "lp@t.com", "password": "pw"}).get_json()
    token = reg["token"]

    paths = client.get("/api/learner/learning-paths").get_json()["learning_paths"]
    learning_path_id = paths[0]["id"]

    enr = client.post(
        "/api/learner/learning-path/enroll",
        json={"learning_path_id": learning_path_id},
        headers=_auth_headers(token),
    ).get_json()["enrollment"]

    all_enr = client.get("/api/admin/enrollments", headers=_auth_headers(admin_token)).get_json()["enrollments"]
    approve_id = next(e["id"] for e in all_enr if e["user_id"] == reg["user"]["id"] and e.get("learning_path_id") == learning_path_id)
    r_app = client.post(f"/api/admin/enrollments/{approve_id}/approve", headers=_auth_headers(admin_token))
    assert r_app.status_code == 200

    outline = client.get(
        f"/api/learner/learning-path/{learning_path_id}/outline",
        headers=_auth_headers(token),
    ).get_json()

    course1 = outline["courses"][0]
    course2 = outline["courses"][1]

    assert course1["lessons"][0]["enabled"] is True
    assert course2["lessons"][0]["enabled"] is False

    token1 = client.post(
        f"/api/learner/learning-path/{learning_path_id}/playback-token",
        json={"course_id": course1["id"], "lesson_id": course1["lessons"][0]["id"]},
        headers=_auth_headers(token),
    ).get_json()["token"]

    # Complete first course lesson
    r_hb = client.post(
        f"/api/learner/learning-path/{learning_path_id}/heartbeat",
        json={"playback_token": token1, "position_seconds": 95.0, "duration_seconds": 100.0},
        headers=_auth_headers(token),
    )
    assert r_hb.status_code == 200

    outline2 = client.get(
        f"/api/learner/learning-path/{learning_path_id}/outline",
        headers=_auth_headers(token),
    ).get_json()

    assert outline2["courses"][1]["lessons"][0]["enabled"] is True


def test_public_outlines_do_not_expose_youtube_ids(client):
    admin_login = client.post("/api/auth/login", json={"email": "sridhar.voleti@gmail.com", "password": "admin123"}).get_json()
    admin_token = admin_login["token"]

    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["learning_path_title", "course_title", "lesson_title", "youtube_id"])
    ws.append(["Path A", "Course 1", "C1-L1", "abc123"])
    ws.append(["Path A", "Course 1", "C1-L2", "def456"])

    buf = BytesIO()
    wb.save(buf)

    r_import = client.post(
        "/api/admin/import/courses",
        headers=_auth_headers(admin_token),
        data={"file": (BytesIO(buf.getvalue()), "courses.xlsx")},
        content_type="multipart/form-data",
    )
    assert r_import.status_code == 200

    courses = client.get("/api/learner/courses").get_json()["courses"]
    course_id = courses[0]["id"]

    public_course = client.get(f"/api/learner/public/course/{course_id}/outline").get_json()
    assert "youtube_id" not in str(public_course)

    paths = client.get("/api/learner/learning-paths").get_json()["learning_paths"]
    learning_path_id = paths[0]["id"]

    public_lp = client.get(f"/api/learner/public/learning-path/{learning_path_id}/outline").get_json()
    assert "youtube_id" not in str(public_lp)


def test_duration_mismatch_is_rejected(client):
    admin_login = client.post("/api/auth/login", json={"email": "sridhar.voleti@gmail.com", "password": "admin123"}).get_json()
    admin_token = admin_login["token"]

    # Import a minimal course
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["course_title", "lesson_title", "youtube_id"])
    ws.append(["Course 1", "L1", "abc123"])

    buf = BytesIO()
    wb.save(buf)

    r_import = client.post(
        "/api/admin/import/courses",
        headers=_auth_headers(admin_token),
        data={"file": (BytesIO(buf.getvalue()), "courses.xlsx")},
        content_type="multipart/form-data",
    )
    assert r_import.status_code == 200

    reg = client.post("/api/auth/register", json={"email": "dm@t.com", "password": "pw"}).get_json()
    token = reg["token"]

    courses = client.get("/api/learner/courses").get_json()["courses"]
    course_id = courses[0]["id"]

    client.post("/api/learner/enroll", json={"course_id": course_id}, headers=_auth_headers(token))

    all_enr = client.get("/api/admin/enrollments", headers=_auth_headers(admin_token)).get_json()["enrollments"]
    approve_id = next(e["id"] for e in all_enr if e["user_id"] == reg["user"]["id"])
    client.post(f"/api/admin/enrollments/{approve_id}/approve", headers=_auth_headers(admin_token))

    outline = client.get(f"/api/learner/course/{course_id}/outline", headers=_auth_headers(token)).get_json()
    lesson_id = outline["lessons"][0]["id"]

    playback = client.post(
        f"/api/learner/course/{course_id}/playback-token",
        json={"lesson_id": lesson_id},
        headers=_auth_headers(token),
    ).get_json()["token"]

    # First heartbeat stores duration = 100
    r1 = client.post(
        f"/api/learner/course/{course_id}/heartbeat",
        json={"playback_token": playback, "position_seconds": 10.0, "duration_seconds": 100.0},
        headers=_auth_headers(token),
    )
    assert r1.status_code == 200

    # Second heartbeat with different duration should be rejected
    r2 = client.post(
        f"/api/learner/course/{course_id}/heartbeat",
        json={"playback_token": playback, "position_seconds": 20.0, "duration_seconds": 130.0},
        headers=_auth_headers(token),
    )
    assert r2.status_code == 400
    assert r2.get_json()["error"] == "duration_mismatch"


def test_me_returns_only_courses_in_active_learning_path(client):
    admin_login = client.post("/api/auth/login", json={"email": "sridhar.voleti@gmail.com", "password": "admin123"}).get_json()
    admin_token = admin_login["token"]

    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["learning_path_title", "course_title", "lesson_title", "youtube_id"])
    ws.append(["Path X", "Course 1", "C1-L1", "abc123"])
    ws.append(["Path X", "Course 2", "C2-L1", "def456"])

    buf = BytesIO()
    wb.save(buf)

    r_import = client.post(
        "/api/admin/import/courses",
        headers=_auth_headers(admin_token),
        data={"file": (BytesIO(buf.getvalue()), "courses.xlsx")},
        content_type="multipart/form-data",
    )
    assert r_import.status_code == 200

    reg = client.post("/api/auth/register", json={"email": "me@t.com", "password": "pw"}).get_json()
    token = reg["token"]

    paths = client.get("/api/learner/learning-paths").get_json()["learning_paths"]
    learning_path_id = paths[0]["id"]

    client.post(
        "/api/learner/learning-path/enroll",
        json={"learning_path_id": learning_path_id},
        headers=_auth_headers(token),
    )

    all_enr = client.get("/api/admin/enrollments", headers=_auth_headers(admin_token)).get_json()["enrollments"]
    approve_id = next(e["id"] for e in all_enr if e["user_id"] == reg["user"]["id"] and e.get("learning_path_id") == learning_path_id)
    client.post(f"/api/admin/enrollments/{approve_id}/approve", headers=_auth_headers(admin_token))

    me = client.get("/api/learner/me", headers=_auth_headers(token)).get_json()
    assert "learning_paths" in me
    assert len(me["learning_paths"]) == 1
    assert me["learning_paths"][0]["id"] == learning_path_id

    courses = me["learning_paths"][0]["courses"]
    assert len(courses) == 2
    assert {c["title"] for c in courses} == {"Course 1", "Course 2"}


def test_course_outline_has_progress_summary(client):
    admin_login = client.post("/api/auth/login", json={"email": "sridhar.voleti@gmail.com", "password": "admin123"}).get_json()
    admin_token = admin_login["token"]

    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["course_title", "lesson_title", "youtube_id"])
    ws.append(["Course P", "L1", "abc123"])
    ws.append(["Course P", "L2", "def456"])

    buf = BytesIO()
    wb.save(buf)

    r_import = client.post(
        "/api/admin/import/courses",
        headers=_auth_headers(admin_token),
        data={"file": (BytesIO(buf.getvalue()), "courses.xlsx")},
        content_type="multipart/form-data",
    )
    assert r_import.status_code == 200

    reg = client.post("/api/auth/register", json={"email": "ps@t.com", "password": "pw"}).get_json()
    token = reg["token"]

    courses = client.get("/api/learner/courses").get_json()["courses"]
    course_id = courses[0]["id"]

    client.post("/api/learner/enroll", json={"course_id": course_id}, headers=_auth_headers(token))

    all_enr = client.get("/api/admin/enrollments", headers=_auth_headers(admin_token)).get_json()["enrollments"]
    approve_id = next(e["id"] for e in all_enr if e["user_id"] == reg["user"]["id"])
    client.post(f"/api/admin/enrollments/{approve_id}/approve", headers=_auth_headers(admin_token))

    outline = client.get(f"/api/learner/course/{course_id}/outline", headers=_auth_headers(token)).get_json()
    assert "progress_summary" in outline
    assert outline["progress_summary"]["total_lessons"] == 2
    assert outline["progress_summary"]["completed_lessons"] == 0
